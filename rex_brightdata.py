"""
Rex Airlines Flight Scraper — Brightdata Edition
=================================================
Scraping logic from rex_new_910.py (TIME FIX v2) — unchanged.
Infrastructure: Brightdata Scraping Browser (Playwright CDP, port 9222).
"""

import asyncio
import os
import re
import sys
import random
import argparse
import requests
from datetime import datetime, timedelta
from pathlib import Path
from zoneinfo import ZoneInfo
from playwright.async_api import async_playwright
from openpyxl import load_workbook, Workbook

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

# ─────────────────────────────────────────────────────────────
#  BRIGHTDATA CREDENTIALS
# ─────────────────────────────────────────────────────────────
BD_BROWSER_HOST = os.getenv("BD_BROWSER_HOST", "brd.superproxy.io")
BD_BROWSER_PORT = int(os.getenv("BD_BROWSER_PORT", "9222"))
BD_BROWSER_USER = os.getenv("BD_BROWSER_USER", "brd-customer-hl_fbc4a16a-zone-cont_rex")
BD_BROWSER_PASS = os.getenv("BD_BROWSER_PASS", "072res2p22t3")

BD_AUTH_TOKEN        = os.getenv("BD_AUTH_TOKEN", "7b1cdf1c-e4e0-4b6c-925b-0121031e6bf7")
BD_WEB_UNLOCKER_ZONE = os.getenv("BD_WEB_UNLOCKER_ZONE", "unblocker1")
BD_UNLOCKER_ENDPOINT = os.getenv("BD_UNLOCKER_ENDPOINT", "https://api.brightdata.com/request")

# ─────────────────────────────────────────────────────────────
#  CONFIGURATION  (unchanged from rex_new_910.py)
# ─────────────────────────────────────────────────────────────

AIRPORT_MAP = {
    "ALH": "Albany",
    "PER": "Perth",
    "EPR": "Esperance",
    "CVQ": "Carnarvon",
    "MJK": "Monkey Mia",
}

ALL_ROUTES = [
    ("PER", "ALH"), ("ALH", "PER"),
    ("PER", "EPR"), ("EPR", "PER"),
    ("PER", "CVQ"), ("CVQ", "PER"),
    ("PER", "MJK"), ("MJK", "PER"),
    ("CVQ", "MJK"), ("MJK", "CVQ"),
]

CONNECTING_ROUTES = {("CVQ", "MJK"), ("MJK", "CVQ")}
NO_RIBBON_ROUTES  = {("PER", "MJK"), ("MJK", "PER")}

TOTAL_DAYS   = 84
OUTPUT_EXCEL = os.getenv("REX_OUTPUT_EXCEL", "output/rex_results_all_routes.xlsx")

CSV_FIELDS = [
    "Date Checked", "Time Checked", "Airline",
    "Date of Departure", "Time of Departure",
    "Origin", "Destination",
    "Fare Price", "Fare Class", "Source",
]

RIBBON_SELECTORS = [
    ".calendar .day",
    ".date-tab",
    ".ribbon-date",
    ".calendar-day",
]

CARD_SELECTORS = [
    ".trip-select .trip",
    ".flight-select-row",
    ".avail-flight-row",
    ".flight-option",
    "div.trip",
    "li.trip",
    ".avail-row",
    "tr.flight-row",
]

FLIGHT_LIST_SELECTORS = [
    ".trip-select",
    ".flight-results",
    ".avail-flights",
    ".departing-block",
    ".flight-list",
    "#flightResults",
]

# ─────────────────────────────────────────────────────────────
#  UTILITIES  (unchanged from rex_new_910.py)
# ─────────────────────────────────────────────────────────────

def today_dt() -> datetime:
    now_au = datetime.now(ZoneInfo("Australia/Sydney")).replace(tzinfo=None)
    return now_au.replace(hour=0, minute=0, second=0, microsecond=0)


def build_date_list() -> list[datetime]:
    t = today_dt()
    return [t + timedelta(days=i) for i in range(TOTAL_DAYS)]


def normalise_time(raw: str) -> str:
    return raw.strip()


def extract_all_times_from_text(text: str) -> list[str]:
    times = re.findall(r'\d{1,2}:\d{2}\s*[aApP][mM]', text)
    if times:
        return [t.strip() for t in times]
    return [t.strip() for t in re.findall(r'\b\d{1,2}:\d{2}\b', text)]


def extract_times_per_zl(card_text: str) -> dict[str, str]:
    all_times  = extract_all_times_from_text(card_text)
    zl_matches = list(re.finditer(r'ZL\s?\d{3,4}', card_text))
    result = {}
    for idx, m in enumerate(zl_matches):
        flight  = re.sub(r'\s', '', m.group())
        dep_idx = idx * 2
        dep     = all_times[dep_idx] if dep_idx < len(all_times) else "-"
        result[flight] = dep
    return result


def _ensure_cents(val: str) -> str:
    return val if '.' in val else f"{val}.00"


def append_rows(rows: list):
    if not rows:
        return
    Path(OUTPUT_EXCEL).parent.mkdir(parents=True, exist_ok=True)
    if os.path.exists(OUTPUT_EXCEL):
        wb = load_workbook(OUTPUT_EXCEL)
        ws = wb.active
    else:
        wb = Workbook()
        ws  = wb.active
        ws.title = "Rex Flight Data"
        ws.append(CSV_FIELDS)
    for row in rows:
        ws.append([row.get(f, "") for f in CSV_FIELDS])
    wb.save(OUTPUT_EXCEL)

# ─────────────────────────────────────────────────────────────
#  PRICE EXTRACTION  (unchanged from rex_new_910.py)
# ─────────────────────────────────────────────────────────────

def extract_price_from_card_text(card_text: str) -> str:
    m = re.search(r'[Ff]rom\s*\$\s*([\d,]+\.\d{2})', card_text)
    if m:
        return f"${m.group(1)}"
    m = re.search(r'\$\s*([\d,]+\.\d{2})', card_text)
    if m:
        return f"${m.group(1)}"
    m = re.search(r'\$\s*(\d[\d,]*)', card_text)
    if m:
        return f"${_ensure_cents(m.group(1))}"
    return "N/A"


def find_ribbon_end_position(full_body: str) -> int:
    markers = [
        "Departure Time",
        "Select your departing flight",
        "departing flight",
        "Fly Economy",
        "Select Fares",
    ]
    for marker in markers:
        pos = full_body.find(marker)
        if pos > 0:
            return pos
    return len(full_body) // 7


def extract_price_from_flight_window(full_body: str, flight_match, ribbon_end_pos: int):
    if flight_match.start() < ribbon_end_pos:
        return None
    after_start = flight_match.end()
    after_end   = min(len(full_body), after_start + 600)
    window      = full_body[after_start:after_end]
    m = re.search(r'[Ff]rom\s*\$\s*([\d,]+\.\d{2})', window)
    if m:
        return f"${m.group(1)}"
    m = re.search(r'[Ss]elect\s+[Ff]ares?\s*\$\s*([\d,]+\.\d{2})', window)
    if m:
        return f"${m.group(1)}"
    decimals = re.findall(r'\$\s*([\d,]+\.\d{2})', window)
    if decimals:
        return f"${decimals[0]}"
    m = re.search(r'\$\s*(\d[\d,]*)', window)
    if m:
        return f"${_ensure_cents(m.group(1))}"
    return "N/A"

# ─────────────────────────────────────────────────────────────
#  BRIGHTDATA HELPERS
# ─────────────────────────────────────────────────────────────

async def wait_for_captcha(page, detect_timeout: int = 60000, solve_timeout: float = 120.0):
    print("   ⏳ Bright Data CAPTCHA solver — waiting up to 120s...")
    try:
        client = await page.context.new_cdp_session(page)
        await asyncio.wait_for(
            client.send("Captcha.waitForSolve", {"detectTimeout": detect_timeout}),
            timeout=solve_timeout,
        )
        print("   ✅ CAPTCHA solved")
    except asyncio.TimeoutError:
        print(f"   ⚠️  CAPTCHA timeout ({solve_timeout}s) — continuing")
    except Exception as e:
        print(f"   ⚠️  CAPTCHA: {e} — continuing")


def check_web_unlocker() -> bool:
    try:
        resp = requests.post(
            BD_UNLOCKER_ENDPOINT,
            json={"zone": BD_WEB_UNLOCKER_ZONE,
                  "url": "https://geo.brdtest.com/welcome.txt?product=unlocker&method=api",
                  "format": "raw"},
            headers={"Authorization": f"Bearer {BD_AUTH_TOKEN}",
                     "Content-Type": "application/json"},
            timeout=60,
        )
        resp.raise_for_status()
        print(f"✅ Web Unlocker check OK: {resp.text.strip()[:180]}")
        return True
    except Exception as e:
        print(f"⚠️  Web Unlocker check failed: {e}")
        return False

# ─────────────────────────────────────────────────────────────
#  SCRAPER CLASS  (logic unchanged from rex_new_910.py)
# ─────────────────────────────────────────────────────────────

class RexScraper:

    def __init__(self):
        self._last_ribbon_price = ""

    def parse_tab_date(self, text: str) -> datetime | None:
        m = re.search(
            r'(?:Mon|Tue|Wed|Thu|Fri|Sat|Sun)[a-z]*\s+(\d{1,2})\s+'
            r'(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)',
            text, re.IGNORECASE
        )
        if not m:
            return None
        day, mon = int(m.group(1)), m.group(2).capitalize()
        now = datetime.now(ZoneInfo("Australia/Sydney")).replace(tzinfo=None)
        for year in [now.year, now.year + 1]:
            try:
                dt = datetime.strptime(f"{day} {mon} {year}", "%d %b %Y")
                if abs((dt - now).days) < 200:
                    return dt
            except ValueError:
                pass
        return None

    async def click_ribbon_tab(self, page, target_dt: datetime) -> str:
        for sel in RIBBON_SELECTORS:
            tabs = await page.query_selector_all(sel)
            if not tabs:
                continue
            for tab in tabs:
                raw = (await tab.inner_text(timeout=3000)).strip().replace("\n", " ")
                tab_dt = self.parse_tab_date(raw)
                if not tab_dt or tab_dt.date() != target_dt.date():
                    continue

                cls       = (await tab.get_attribute("class") or "").lower()
                raw_lower = raw.lower()
                no_flight = any([
                    "unavailable" in raw_lower,
                    "unavailable" in cls,
                    "disabled"    in cls,
                    "no-flight"   in cls,
                    "noflight"    in cls,
                    "greyed"      in cls,
                    "inactive"    in cls,
                    "$" not in raw,
                ])
                if no_flight:
                    print(f"   ℹ️  No-flight tab: '{raw[:70]}'")
                    self._last_ribbon_price = ""
                    return "unavailable"

                rp = re.search(r'\$\s*([\d,]+(?:\.\d{2})?)', raw)
                if rp:
                    val = rp.group(1)
                    self._last_ribbon_price = f"${_ensure_cents(val)}"
                    print(f"   🎫 Ribbon price: {self._last_ribbon_price}")

                await tab.click(force=True)
                return "clicked"
        return "not_found"

    async def go_next_ribbon(self, page) -> bool:
        for sel in [
            ".calendar .arrow.next", ".ribbon-next", "button.next-week",
            "[aria-label='Next week']", ".date-nav-next",
            "span.arrow.right", "button[class*='next']",
            "span[class*='arrow']", "a[class*='next']",
        ]:
            btn = await page.query_selector(sel)
            if btn:
                if (await btn.get_attribute("disabled")) is not None:
                    return False
                cls = (await btn.get_attribute("class") or "").lower()
                if "disabled" in cls:
                    return False
                await btn.click()
                await asyncio.sleep(3)
                return True
        return False

    async def wait_for_flights_loaded(self, page, target_dt: datetime,
                                       timeout: int = 15) -> bool:
        exp_day    = str(target_dt.day)
        exp_mon    = target_dt.strftime("%b")
        date_ok    = False
        flights_ok = False

        for _ in range(timeout * 2):
            if not date_ok:
                for sel in [
                    ".departing-block h2.date", ".selected-date",
                    "h2.date", ".flight-date-header", ".date-heading",
                    "h2", ".date-display",
                ]:
                    try:
                        txt = await page.locator(sel).first.inner_text(timeout=500)
                        if exp_day in txt and exp_mon in txt:
                            date_ok = True
                            break
                    except Exception:
                        pass

            if not flights_ok:
                for sel in CARD_SELECTORS:
                    cards = await page.query_selector_all(sel)
                    if cards:
                        for card in cards[:3]:
                            txt = (await card.inner_text(timeout=3000)).strip()
                            if re.search(r'ZL\s?\d{3,4}', txt):
                                flights_ok = True
                                break
                    if flights_ok:
                        break

                if not flights_ok:
                    try:
                        body_snippet = await page.inner_text("body", timeout=5000)
                        if re.search(r'ZL\s?\d{3,4}', body_snippet[:5000]):
                            flights_ok = True
                    except Exception:
                        pass

            if date_ok and flights_ok:
                print("   ✅ Page loaded — date synced + flights visible")
                return True

            await asyncio.sleep(0.5)

        if flights_ok:
            print("   ⚠️  Flights visible but date header uncertain — proceeding")
            return True

        print(f"   ⚠️  Load timeout ({timeout}s) — extracting anyway")
        return False

    # ── Extract flights ──────────────────────────────────────────
    async def extract_flights(self, page, date_str, origin, dest) -> list:
        now        = datetime.now(ZoneInfo("Australia/Sydney")).replace(tzinfo=None)
        ck_date    = now.strftime("%d-%m-%Y")
        ck_time    = now.strftime("%H:%M:%S")
        data, seen = [], set()
        connecting = (origin, dest) in CONNECTING_ROUTES

        # ── STEP 1: Card selectors ───────────────────────────────
        for sel in CARD_SELECTORS:
            rows = await page.query_selector_all(sel)
            if not rows:
                continue
            print(f"   🔍 Card selector '{sel}' → {len(rows)} row(s)")

            for row in rows:
                try:
                    text = (await row.inner_text(timeout=5000)).strip()
                except Exception:
                    continue
                flat        = text.replace("\n", " ")
                zl_time_map = extract_times_per_zl(text)
                if not zl_time_map:
                    continue

                price = extract_price_from_card_text(flat)

                for f_no, dep in zl_time_map.items():
                    print(f"      Card: {f_no}  dep={dep}  price={price}")
                    key = f"{f_no}-{dep}"
                    if key not in seen:
                        data.append(self._row(ck_date, ck_time, f_no,
                                              date_str, dep, origin, dest, price))
                        seen.add(key)

            if data:
                print(f"   ✅ Card extraction done: {len(data)} flight(s)")
                return data

        # ── STEP 2: Body text scan (fallback) ───────────────────
        print("   ⚠️ No card selector matched — body text scan (ribbon-aware).")
        try:
            full_body = await page.inner_text("body", timeout=8000)
        except Exception:
            return data

        ribbon_end = find_ribbon_end_position(full_body)
        print(f"   📍 Ribbon area ends at ~char {ribbon_end}")

        zl_matches = list(re.finditer(r'ZL\s?\d{3,4}', full_body))
        print(f"   🔍 Total ZL matches: {len(zl_matches)} (ribbon cutoff pos: {ribbon_end})")

        flights_raw       = []
        body_after_ribbon = full_body[ribbon_end:]
        all_body_times    = extract_all_times_from_text(body_after_ribbon)
        zl_in_body        = [m for m in zl_matches if m.start() >= ribbon_end]

        for idx, m in enumerate(zl_in_body):
            f_no    = re.sub(r'\s', '', m.group())
            dep_idx = idx * 2
            dep     = all_body_times[dep_idx] if dep_idx < len(all_body_times) else "-"
            key     = f"{f_no}-{dep}"
            if key in seen:
                continue
            price = extract_price_from_flight_window(full_body, m, ribbon_end)
            if price is None:
                continue
            flights_raw.append((f_no, dep, key, price))
            seen.add(key)

        if not flights_raw:
            return data

        print(f"   🔍 Body scan (post-ribbon) found: {[f[0] for f in flights_raw]}")

        if connecting:
            flight_area    = full_body[ribbon_end:]
            m_fp           = re.search(r'[Ff]rom\s*\$\s*([\d,]+\.\d{2})', flight_area)
            combined_price = (f"${m_fp.group(1)}" if m_fp
                              else self._last_ribbon_price or "N/A")
            print(f"   🔗 Connecting route — combined price: {combined_price}")
            for f_no, dep, key, _ in flights_raw:
                data.append(self._row(ck_date, ck_time, f_no,
                                      date_str, dep, origin, dest, combined_price))
        else:
            for f_no, dep, key, price in flights_raw:
                print(f"      ✈️  {f_no}  {dep} → {price}")
                data.append(self._row(ck_date, ck_time, f_no,
                                      date_str, dep, origin, dest, price))

        return data

    def _row(self, ck_date, ck_time, airline, dep_date, dep_time, orig, dest, price):
        return {
            "Date Checked":      ck_date,
            "Time Checked":      ck_time,
            "Airline":           airline,
            "Date of Departure": dep_date,
            "Time of Departure": dep_time,
            "Origin":            orig,
            "Destination":       dest,
            "Fare Price":        price,
            "Fare Class":        "Economy",
            "Source":            "Rex Website",
        }

    def _no_flight(self, date_str, orig, dest):
        now = datetime.now(ZoneInfo("Australia/Sydney")).replace(tzinfo=None)
        return self._row(now.strftime("%d-%m-%Y"), now.strftime("%H:%M:%S"),
                         "no flight", date_str, "-", orig, dest, "-")

    # ─────────────────────────────────────────────────────────
    #  FRESH SEARCH  (unchanged logic from rex_new_910.py)
    # ─────────────────────────────────────────────────────────
    async def do_fresh_search(self, page, origin_name: str, dest_name: str,
                               target_dt: datetime) -> bool:
        print(f"   🔄 Fresh search: {origin_name} → {dest_name} "
              f"on {target_dt.strftime('%d %b %Y')}")
        try:
            await page.goto("https://www.rex.com.au/")
            await wait_for_captcha(page)
            try:
                await page.get_by_role("button", name="Continue").click(timeout=3000)
            except Exception:
                pass

            await page.locator("label[for*='rbTripType_oneway']").click()

            await page.locator(
                "#ContentPlaceHolder1_BookingHomepageV21_OriginAirport + .select2-container"
            ).click()
            await page.locator(".select2-search__field").fill(origin_name)
            await page.locator(".select2-results__option").filter(
                has_text=origin_name
            ).first.click()

            await page.locator(
                "#ContentPlaceHolder1_BookingHomepageV21_DestinationAirport + .select2-container"
            ).click()
            await page.locator(".select2-search__field").fill(dest_name)
            await page.locator(".select2-results__option").filter(
                has_text=dest_name
            ).first.click()

            await page.locator("#datefilter").click()
            await asyncio.sleep(1)

            for _ in range(12):
                try:
                    month_txt = await page.locator(
                        ".daterangepicker .month"
                    ).first.inner_text(timeout=1000)
                    visible_month = datetime.strptime(month_txt.strip(), "%b %Y")
                    if (visible_month.year  == target_dt.year and
                            visible_month.month == target_dt.month):
                        break
                    await page.locator(".daterangepicker .next").first.click()
                    await asyncio.sleep(0.5)
                except Exception:
                    break

            await page.locator(
                ".daterangepicker td.available:not(.off)"
            ).filter(has_text=str(target_dt.day)).first.click()
            await asyncio.sleep(0.5)

            await page.locator(
                "#ContentPlaceHolder1_BookingHomepageV21_SubmitBooking"
            ).click()

            await wait_for_captcha(page)
            loaded = await self.wait_for_flights_loaded(page, target_dt, timeout=20)
            await asyncio.sleep(1)
            print(f"   ✅ Fresh search complete — loaded={loaded}")
            return True

        except Exception as e:
            print(f"   ❌ Fresh search failed: {e}")
            return False

    # ─────────────────────────────────────────────────────────
    #  RIBBON ROUTE  (unchanged logic from rex_new_910.py)
    # ─────────────────────────────────────────────────────────
    async def run_ribbon_route(self, page, origin_code, dest_code):
        all_dates   = build_date_list()
        origin_name = AIRPORT_MAP.get(origin_code, origin_code)
        dest_name   = AIRPORT_MAP.get(dest_code, dest_code)

        for idx, target_dt in enumerate(all_dates, 1):
            date_str = target_dt.strftime("%d-%m-%Y")

            print(f"\n{'═'*60}")
            print(f"📅 [{idx}/{TOTAL_DAYS}]  {target_dt.strftime('%A, %d %b %Y')}")
            print(f"{'─'*60}")

            self._last_ribbon_price = ""

            tab_result = "not_found"
            for attempt in range(30):
                tab_result = await self.click_ribbon_tab(page, target_dt)
                if tab_result in ("clicked", "unavailable"):
                    break
                print(f"   ➡️  Not in view — advancing ribbon (attempt {attempt+1})...")
                moved = await self.go_next_ribbon(page)
                if not moved:
                    print("   ⛔ Ribbon end reached.")
                    tab_result = "ribbon_end"
                    break
                await asyncio.sleep(3)
            else:
                tab_result = "not_found"

            if tab_result in ("not_found", "ribbon_end"):
                print("   🔄 Ribbon miss — fresh search...")
                ok = await self.do_fresh_search(page, origin_name, dest_name, target_dt)
                if not ok:
                    print("   ❌ Fresh search fail — 'no flight'")
                    append_rows([self._no_flight(date_str, origin_code, dest_code)])
                    continue
                tab_result_after = await self.click_ribbon_tab(page, target_dt)
                if tab_result_after == "unavailable":
                    print("   ❌ Unavailable after fresh search — 'no flight'")
                    append_rows([self._no_flight(date_str, origin_code, dest_code)])
                    continue
                print(f"   ℹ️  After fresh search tab_result={tab_result_after} — extracting")

            elif tab_result == "unavailable":
                print("   ❌ No flight → 'no flight'")
                append_rows([self._no_flight(date_str, origin_code, dest_code)])
                continue

            await self.wait_for_flights_loaded(page, target_dt, timeout=15)
            await asyncio.sleep(1)

            flights = await self.extract_flights(page, date_str, origin_code, dest_code)
            if flights:
                print(f"   ✅ {len(flights)} flight(s):")
                for f in flights:
                    print(f"      ✈️  {f['Airline']}  {f['Time of Departure']}  {f['Fare Price']}")
                append_rows(flights)
            else:
                print("   ⚠️  0 found → 'no flight'")
                append_rows([self._no_flight(date_str, origin_code, dest_code)])

            print(f"   ✅ {idx}/{TOTAL_DAYS} done.")

    # ─────────────────────────────────────────────────────────
    #  FRESH SEARCH ROUTE (PER↔MJK — no ribbon)
    # ─────────────────────────────────────────────────────────
    async def run_fresh_search_route(self, page, origin_code, dest_code):
        all_dates   = build_date_list()
        origin_name = AIRPORT_MAP.get(origin_code, origin_code)
        dest_name   = AIRPORT_MAP.get(dest_code, dest_code)

        for idx, target_dt in enumerate(all_dates, 1):
            date_str = target_dt.strftime("%d-%m-%Y")

            print(f"\n{'═'*60}")
            print(f"📅 [{idx}/{TOTAL_DAYS}]  {target_dt.strftime('%A, %d %b %Y')}")
            print(f"{'─'*60}")

            self._last_ribbon_price = ""

            ok = await self.do_fresh_search(page, origin_name, dest_name, target_dt)
            if not ok:
                print("   ❌ Fresh search fail — 'no flight'")
                append_rows([self._no_flight(date_str, origin_code, dest_code)])
                continue

            flights = await self.extract_flights(page, date_str, origin_code, dest_code)
            if flights:
                print(f"   ✅ {len(flights)} flight(s):")
                for f in flights:
                    print(f"      ✈️  {f['Airline']}  {f['Time of Departure']}  {f['Fare Price']}")
                append_rows(flights)
            else:
                print("   ⚠️  0 found — 'no flight'")
                append_rows([self._no_flight(date_str, origin_code, dest_code)])

            print(f"   ✅ {idx}/{TOTAL_DAYS} done.")

    # ─────────────────────────────────────────────────────────
    #  ROUTE ENTRY POINT
    # ─────────────────────────────────────────────────────────
    async def run_route(self, origin_code, dest_code):
        origin_name = AIRPORT_MAP.get(origin_code, origin_code)
        dest_name   = AIRPORT_MAP.get(dest_code, dest_code)
        dates       = build_date_list()

        print(f"\n{'█'*60}")
        print(f"  ROUTE : {origin_code} ({origin_name}) → {dest_code} ({dest_name})")
        rtype = "Connecting" if (origin_code, dest_code) in CONNECTING_ROUTES else "Independent"
        print(f"  Type  : {rtype}")
        print(f"  Window: {dates[0].strftime('%d-%m-%Y')} → {dates[-1].strftime('%d-%m-%Y')}")
        print(f"  Output: {OUTPUT_EXCEL}")
        print(f"{'█'*60}")

        sid        = random.randint(1000000, 9999999)
        user       = f"{BD_BROWSER_USER}-session-{sid}"
        sbr_ws_cdp = f"wss://{user}:{BD_BROWSER_PASS}@{BD_BROWSER_HOST}:{BD_BROWSER_PORT}"

        print(f"🔌 Connecting to Bright Data Scraping Browser (Playwright CDP)...")
        print(f"   Session: {sid}")

        async with async_playwright() as p:
            try:
                browser = await p.chromium.connect_over_cdp(sbr_ws_cdp)
            except Exception as e:
                print(f"❌ Bright Data connection failed: {e}")
                append_rows([{
                    "Date Checked":      datetime.now().strftime("%d-%m-%Y"),
                    "Time Checked":      datetime.now().strftime("%H:%M:%S"),
                    "Airline":           "site unavailable",
                    "Date of Departure": dt.strftime("%d-%m-%Y"),
                    "Time of Departure": "-",
                    "Origin":            origin_code,
                    "Destination":       dest_code,
                    "Fare Price":        "-",
                    "Fare Class":        "Economy",
                    "Source":            "Rex Website - unavailable",
                } for dt in dates])
                return

            print("✅ Connected to Bright Data!")
            page = await browser.new_page()

            try:
                await page.goto("https://www.rex.com.au/")
                await wait_for_captcha(page)
                try:
                    await page.get_by_role("button", name="Continue").click(timeout=3000)
                except Exception:
                    pass

                await page.locator("label[for*='rbTripType_oneway']").click()

                await page.locator(
                    "#ContentPlaceHolder1_BookingHomepageV21_OriginAirport + .select2-container"
                ).click()
                await page.locator(".select2-search__field").fill(origin_name)
                await page.locator(".select2-results__option").filter(
                    has_text=origin_name
                ).first.click()

                await page.locator(
                    "#ContentPlaceHolder1_BookingHomepageV21_DestinationAirport + .select2-container"
                ).click()
                await page.locator(".select2-search__field").fill(dest_name)
                await page.locator(".select2-results__option").filter(
                    has_text=dest_name
                ).first.click()

                today = datetime.now(ZoneInfo("Australia/Sydney")).replace(tzinfo=None)
                await page.locator("#datefilter").click()
                await page.locator(
                    ".daterangepicker td.available:not(.off)"
                ).filter(has_text=str(today.day)).first.click()

                await page.locator(
                    "#ContentPlaceHolder1_BookingHomepageV21_SubmitBooking"
                ).click()

                print("⏳ Waiting for flight page + CAPTCHA solve...")
                await wait_for_captcha(page)

                if (origin_code, dest_code) in NO_RIBBON_ROUTES:
                    await self.run_fresh_search_route(page, origin_code, dest_code)
                else:
                    await self.run_ribbon_route(page, origin_code, dest_code)

            except KeyboardInterrupt:
                print("\n⛔ Interrupted.")
            except Exception as e:
                print(f"\n❌ Fatal error in run_route({origin_code}→{dest_code}): {e}")
                import traceback
                traceback.print_exc()
            finally:
                try:
                    await browser.close()
                except Exception:
                    pass

        print(f"\n  📊 Output saved: {OUTPUT_EXCEL}\n")


# ─────────────────────────────────────────────────────────────
#  ENTRY POINT
# ─────────────────────────────────────────────────────────────

def main():
    global OUTPUT_EXCEL
    parser = argparse.ArgumentParser(description="Rex Airlines Scraper (Brightdata)")
    parser.add_argument("--skip-unblocker-check", action="store_true")
    parser.add_argument("--output", default=OUTPUT_EXCEL)
    args = parser.parse_args()

    OUTPUT_EXCEL = args.output
    Path(OUTPUT_EXCEL).parent.mkdir(parents=True, exist_ok=True)

    today = today_dt()
    dates = build_date_list()
    print("🗓️  TODAY =", today.strftime("%A, %d-%m-%Y"), " (Day 0)")
    print(f"📋 Routes: {len(ALL_ROUTES)}")
    print(f"📆 Dates per route: {TOTAL_DAYS}")
    print(f"🌐 Bright Data: {BD_BROWSER_HOST}:{BD_BROWSER_PORT} (Playwright CDP)")
    print(f"📊 Output: {OUTPUT_EXCEL}")
    print("█" * 60)

    if not args.skip_unblocker_check:
        check_web_unlocker()

    scraper = RexScraper()
    for origin, dest in ALL_ROUTES:
        try:
            asyncio.run(scraper.run_route(origin, dest))
        except KeyboardInterrupt:
            print(f"\n⛔ Stopped at {origin}→{dest}.")
            break
        import time
        time.sleep(5)

    print("\n✅ All routes complete.")


if __name__ == "__main__":
    main()
